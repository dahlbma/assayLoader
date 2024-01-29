from scipy.optimize import curve_fit
import os
import json
import sys
import numpy as np
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
from PyQt5.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font


import warnings
warnings.filterwarnings('ignore')


# Define the 4-PL model function
def fourpl(x, slope, ic50, bottom, top):
    return bottom + (top - bottom) / (1 + (x / ic50)**slope)

file_path = "finalPreparedDR.xlsx"

class ScatterplotWidget(QWidget):
    def __init__(self, data_dict, rowPosition, parent=None):
        super(ScatterplotWidget, self).__init__(parent)
        self.rowPosition = rowPosition

        self.figure, self.ax = plt.subplots(figsize=(3, 2))
        self.canvas = FigureCanvas(self.figure)

        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.setLayout(layout)
        slope, ic50, bottom, top = self.plot_scatter(data_dict)
        self.ic50 = ic50
        self.slope = slope
        self.top = top
        self.bottom = bottom
        self.minConc = data_dict['finalConc_nL'].iloc[0]
        self.maxConc = data_dict['finalConc_nL'].iloc[-1]

    def plot_scatter(self, df):

        # Extract the 'x' and 'y' arrays
        x_values = np.array(df['finalConc_nL'].values/1000000000, dtype=np.float64)
        #y_values = np.array(df['yMean'].values, dtype=np.float64)
        y_values = np.array(df['inhibition'].values, dtype=np.float64)
        y_err_values = np.array(df['yStd'].values, dtype=np.float64)
        
        fitOk = True
        try:
            top = np.max(y_values)
            bottom = np.min(y_values)
            slope = -1
            ic50 = np.mean(x_values)*2
            top = 100
            bottom = 0
            # Fit the data to the 4-PL model
            params, covariance = curve_fit(fourpl,
                                           x_values,
                                           y_values,
                                           maxfev = 10000,
                                           p0=[slope, ic50, bottom, top],
                                           bounds=([-100, 0, -20, 70], [10, 0.01, 40, 120])
                                           )
            #print(f'covariance: {covariance}')
            # Extract the fitted parameters
            slope, ic50, bottom, top = params
        except Exception as e:
            print(f'''Can't fit parameters {str(e)} {x_values[0]}''')
            fitOk = False
            slope = -1
            ic50 = -1
            bottom = -1
            top = -1

        # Generate a curve using the fitted parameters
        x_curve = np.logspace(np.log10(min(x_values)), np.log10(max(x_values)), 100)
        if fitOk == True:
            y_curve_fit = fourpl(x_curve, slope, ic50, bottom, top)

        # Plot the original data and the fitted curve with a logarithmic x-axis
        #plt.scatter(x_values, y_values, label='Original Data')

        plt.errorbar(x_values, y_values, yerr=y_err_values, fmt='o', label='Raw data')

        plt.ylim(min(min(y_values), 0) - 10, max(max(y_values), 100) + 10)
        
        if fitOk == True:
            plt.plot(x_curve, y_curve_fit, label='Fitted 4-PL Curve')
        if ic50 == -1:
            pass
        elif ic50 > 0.001:
            plt.axvline(ic50, color='r', linestyle='--', label=f'IC50 = {ic50:.2f} M')
        else:
            plt.axvline(ic50, color='r', linestyle='--', label=f'IC50 = {ic50*1e6:.2f} uM')
        plt.xscale('log')  # Set x-axis to logarithmic scale
        plt.xlabel('Concentration')
        plt.ylabel('Inhibition %')
        self.canvas.draw()
        self.figure.savefig(f'img/{self.rowPosition}.png', bbox_inches='tight', dpi=96)
        plt.legend()
        self.canvas.draw()

        return slope, ic50, bottom, top


class ScatterplotsTable(QMainWindow):
    def __init__(self):
        super(ScatterplotsTable, self).__init__()

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        self.table_widget = QTableWidget(self.central_widget)
        # 0            1        2     3    4        5        6
        # Compound_Id, Batch_ID SLOPE IC50 Min.Conc Max.Conc Graph

        
        self.table_widget.setColumnCount(9)
        #self.table_widget.setRowCount(200)

        #for i in range(200):
        #    self.table_widget.setRowHeight(i, 450)

        # Set column width to 800 pixels
        self.table_widget.setColumnWidth(8, 600)

        layout = QVBoxLayout(self.central_widget)
        layout.addWidget(self.table_widget)

        self.generate_scatterplots()
        self.table_widget.setHorizontalHeaderLabels(['Batch', 'Compound', 'IC50', 'Slope', 'Bottom',
                                                     'Top', 'Min Conc nM', 'Max Conc nM', 'Graph'])


    def generate_scatterplots(self):
        df = pd.read_excel(file_path)
        for batch_nr, batch_df in df.groupby('Batch nr'):
            
            # "Batch nr" "Compound ID" "finalConc_nL" "yMean" "yStd"
            batch = batch_df['Batch nr'].iloc[0]
            compound = batch_df['Compound ID'].iloc[0]
            
            rowPosition = self.table_widget.rowCount()


            if rowPosition > 200:
                continue

            
            self.table_widget.insertRow(rowPosition)
            self.table_widget.setRowHeight(rowPosition, 450)
            
            print(f'Plot nr: {rowPosition}')
            # Call the scatter function for each batch
            scatterplot_widget = ScatterplotWidget(batch_df, rowPosition)

            item = QTableWidgetItem(batch)
            self.table_widget.setItem(rowPosition, 0, item)

            item = QTableWidgetItem(compound)
            self.table_widget.setItem(rowPosition, 1, item)

            item = QTableWidgetItem(str("{:.2e}".format(scatterplot_widget.ic50)))
            self.table_widget.setItem(rowPosition, 2, item)

            item = QTableWidgetItem(str(f"{abs(scatterplot_widget.slope):.2f}"))
            self.table_widget.setItem(rowPosition, 3, item)

            item = QTableWidgetItem(str(f"{scatterplot_widget.bottom:.2f}"))
            self.table_widget.setItem(rowPosition, 4, item)

            item = QTableWidgetItem(str(f"{scatterplot_widget.top:.2f}"))
            self.table_widget.setItem(rowPosition, 5, item)            

            item = QTableWidgetItem(str(f"{scatterplot_widget.minConc:.1f}"))
            self.table_widget.setItem(rowPosition, 6, item)

            item = QTableWidgetItem(str(f"{scatterplot_widget.maxConc:.1f}"))
            self.table_widget.setItem(rowPosition, 7, item)

            item = QTableWidgetItem()
            #item.setSizeHint(scatterplot_widget.sizeHint())
            self.table_widget.setItem(rowPosition, 8, item)
            self.table_widget.setCellWidget(rowPosition, 8, scatterplot_widget)
        self.saveToExcel()
        

    def saveToExcel(self):
        # Convert QTableWidget data to a pandas DataFrame
        table_data = []
        for row in range(self.table_widget.rowCount()):
            row_data = [self.table_widget.item(row, col).text() if col < 8 else None for col in range(self.table_widget.columnCount())]
            table_data.append(row_data)

        columns = ['Batch', 'Compound', 'IC50', 'Slope', 'Bottom', 'Top', 'Min Conc nM', 'Max Conc nM', 'Graph']
        df = pd.DataFrame(table_data, columns=columns)

        wb = Workbook()
        ws = wb.active

        file_path = 'DR_Excel.xlsx'


        headings = ["Batch", "Compound", "IC50", "Slope", "Bottom", "Top", "MinConc nM", "MaxConc nM", "Graph"]

        for col_num, heading in enumerate(headings, 1):
            cell = ws.cell(row=1, column=col_num, value=heading)
            cell.font = Font(bold=True)

    
        # Write DataFrame to Excel
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start from row 2 to leave space for header
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

                
        # Add scatterplots to Excel
        for i, canvas in enumerate(df['Graph']):
            image_path = f"img/{i}.png"
            img = Image(image_path)
                        
            ws.add_image(img)
            ws.add_image(img, f"I{i + 2}")

        # Adjust row heights and column widths to fit images
        for r_idx in range(2, len(df) + 2):  # Include header row
            ws.row_dimensions[r_idx].height = 160  # Adjust the height as needed
            
        for c_idx in range(1, len(columns) + 1):
            ws.column_dimensions[chr(ord('A') + c_idx - 1)].width = 13  # Adjust the width as needed

        ws.column_dimensions[chr(ord('I'))].width = 40
        # Save the Excel workbook

        wb.save(file_path)











        '''
        if file_path:
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                for i in range(len(df['Graph'])):
                    #worksheet.insert_image(i + 1, len(columns), f"scatter_{i}.png", {'x_scale': 0.5, 'y_scale': 0.5})
                    worksheet.insert_image(f'I{i + 1}', f"img/{i}.png", {'x_scale': 0.5, 'y_scale': 0.5})
        '''



if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ScatterplotsTable()
    window.setGeometry(100, 100, 350, 1000)  # Adjust window width and height
    window.show()
    sys.exit(app.exec_())
