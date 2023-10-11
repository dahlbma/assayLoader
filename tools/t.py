import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font, NamedStyle
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import numpy as np




# Sample DataFrame (replace with your data)
data = {
    "well": [f"{row}{col:02d}" for row in "ABCDEFGHIJKLMNOP" for col in range(1, 25)],
    "avgValue": [3.5, 4.2, 2.8, 2.0, 2.5, 2.9, 3.7, 4.4, 2.1, 1.8, 2.9, 2.5, 3.2, 3.9, 2.2, 1.9, 2.7, 3.1, 3.3, 2.7, 2.3, 3.1, 2.9, 1.8] * 16
}

df_avg_well = pd.DataFrame(data)

# Calculate average for each column (1-24)
avg_columns = df_avg_well.groupby(df_avg_well["well"].str[1:]).agg({"avgValue": "mean"}).reset_index()
avg_columns.rename(columns={"avgValue": "Avg_Column"}, inplace=True)

# Calculate average for each row (A-P)
avg_rows = df_avg_well.groupby(df_avg_well["well"].str[0]).agg({"avgValue": "mean"}).reset_index()
avg_rows.rename(columns={"avgValue": "Avg_Row"}, inplace=True)

# Calculate standard deviation for each column (1-24)
std_columns = df_avg_well.groupby(df_avg_well["well"].str[1:]).agg({"avgValue": "std"}).reset_index()
std_columns.rename(columns={"avgValue": "Std_Column"}, inplace=True)

# Calculate standard deviation for each row (A-P)
std_rows = df_avg_well.groupby(df_avg_well["well"].str[0]).agg({"avgValue": "std"}).reset_index()
std_rows.rename(columns={"avgValue": "Std_Row"}, inplace=True)

# Display the resulting DataFrames
print("Average for each column (1-24):")
print(avg_columns)

print("\nAverage for each row (A-P):")
print(avg_rows)

print("\nStandard deviation for each column (1-24):")
print(std_columns)

print("\nStandard deviation for each row (A-P):")
print(std_rows)


quit()




# Sample DataFrame (replace with your data)
data = {
    "inhibition": [0.5, None, 0.7, 0.8, 0.9],
    "negCtrlInhibition": [0.4, 0.5, 0.6, 0.7, 0.8],
    "posCtrlInhibition": [0.6, 0.7, 0.8, 0.9, 1.0],
}

df = pd.DataFrame(data)

# Create a scatter plot
plt.scatter(df.index, df["inhibition"], c='blue', label='inhibition')
plt.scatter(df.index, df["negCtrlInhibition"], c='red', label='negCtrlInhibition')
plt.scatter(df.index, df["posCtrlInhibition"], c='red', label='posCtrlInhibition')

# Add labels and legend
plt.xlabel("Index")
plt.ylabel("Inhibition Values")
plt.title("Scatter Plot of Inhibition")
plt.legend()

# Show the plot
plt.show()



quit()



# Create a sample Pandas DataFrame (replace this with your actual DataFrame)
data = {'inhibition': [-2.5, 1.2, 2.3, 1.8, 3.5, 4.2, 2.7, 1.5, 3.8, 4.0, 3.2, 2.0, -1]}
df = pd.DataFrame(data)

# Convert the "inhibition" column to integers (rounded)
df['inhibition'] = df['inhibition'].round().astype(int)

# Plot a histogram of the "inhibition" column
plt.hist(df['inhibition'], bins=10, edgecolor='black')
plt.xlabel('Inhibition (integer)')
plt.ylabel('Frequency')
plt.title('Distribution of Inhibition as Integers')
plt.show()

quit()

# Create a sample DataFrame
data = {'Column1': [1.2345678, 2.3456789, 3.4567890],
        'Column2': [4.5678901, 5.6789012, 6.7890123]}

df = pd.DataFrame(data)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Define a NamedStyle for 'Calibri' font and 2 decimal places
style = NamedStyle(name='custom_style')
style.font = Font(name='Calibri')
style.number_format = '0.00'

for col in 'AB':
     ws._styles[col] = style



# Apply the style to the whole worksheet
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.style = style

# Write the DataFrame to the Excel sheet
for row_data in dataframe_to_rows(df, index=False, header=True):
    ws.append(row_data)

# Save the workbook to a file
excel_file_path = 'output.xlsx'  # Change this to your desired file path
wb.save(excel_file_path)

print(f'DataFrame written to Excel sheet with "Calibri" font and 2 decimal places: {excel_file_path}')

quit()




outFile = 'colored_cells.xlsx'
# Define a custom font (Calibri 10pt)
custom_font = Font(name='Calibri', size=11)

def create_plate_frame(ws, plate_id, top_left_cell, num_columns, num_rows):
    """
    Create an Excel sheet with plate_id, letters A-Z, and numbers 1-num_columns.

    Parameters:
    - ws: The Excel worksheet.
    - plate_id: The text to write in the top_left_cell.
    - top_left_cell: The top-left cell (e.g., 'A1').
    - num_columns: The number of columns for numbers (1-num_columns).
    - num_rows: The number of rows for letters (A-Z).
    """
    
    # Generate letters 'A' to 'Z' below the top_left_cell
    start_row, start_col = ws[top_left_cell].row, ws[top_left_cell].column
    
    # Write the plate_id to the specified cell
    ws[top_left_cell] = "Plate"
    cell = ws.cell(row=start_row, column=start_col+1)
    cell.value = plate_id

    for i in range(num_rows):
        letter = chr(ord('A') + i)
        cell = ws.cell(row=start_row + i + 3, column=start_col)
        cell.value = letter
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Generate numbers 1 to num_columns in a row, one column to the right
    for i in range(num_columns):
        number = i + 1
        cell = ws.cell(row=start_row + 2, column=start_col + i + 1)
        cell.value = number
        cell.alignment = Alignment(horizontal='center', vertical='center')


def create_outer_thick_border(ws, top_left_cell, num_columns, num_rows):
    """
    Create a thick border around the outer border of a block of cells in an Excel worksheet.

    Parameters:
    - ws: The Excel worksheet.
    - top_left_cell: The top-left cell (e.g., 'A1') of the block.
    - num_columns: The number of columns in the block.
    - num_rows: The number of rows in the block.
    """

    top_left_style = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style=None)
    )

    top_right_style = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style=None)
    )

    bottom_left_style = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thick')
    )

    bottom_right_style = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style='thick')
    )

    # Convert the top-left cell to row and column indices
    start_row, start_col = ws[top_left_cell].row + 3, ws[top_left_cell].column + 1

    top_l_cell = ws.cell(row=start_row, column=start_col)
    top_r_cell = ws.cell(row=start_row, column=start_col+num_columns-1)
    bottom_l_cell = ws.cell(row=start_row+num_rows-1, column=start_col)
    bottom_r_cell = ws.cell(row=start_row+num_rows-1, column=start_col+num_columns-1)

    # Apply thick border to top row
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=start_row, column=col)
        cell.border = Border(top=Side(style='thick'))
        top_left_cell = cell

    # Apply thick border to bottom row
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=start_row + num_rows - 1, column=col)
        cell.border = Border(bottom=Side(style='thick'))

    # Apply thick border to left column
    for row in range(start_row, start_row + num_rows):
        cell = ws.cell(row=row, column=start_col)
        cell.border = Border(left=Side(style='thick'))

    # Apply thick border to right column
    for row in range(start_row, start_row + num_rows):
        cell = ws.cell(row=row, column=start_col + num_columns - 1)
        cell.border = Border(right=Side(style='thick'))

    # Apply borders to the corner cells
    top_l_cell.border = top_left_style
    top_r_cell.border = top_right_style
    bottom_l_cell.border = bottom_left_style
    bottom_r_cell.border = bottom_right_style

def generate_gradient(end_color, start_color, num_steps):
    # Parse the start and end colors into RGB components
    start_r = int(start_color[1:3], 16)
    start_g = int(start_color[3:5], 16)
    start_b = int(start_color[5:7], 16)

    end_r = int(end_color[1:3], 16)
    end_g = int(end_color[3:5], 16)
    end_b = int(end_color[5:7], 16)

    # Calculate the step size for each RGB component
    r_step = (end_r - start_r) / num_steps
    g_step = (end_g - start_g) / num_steps
    b_step = (end_b - start_b) / num_steps

    # Generate the gradient colors and store them in a list
    gradient = []
    for i in range(num_steps + 1):
        r = int(start_r + i * r_step)
        g = int(start_g + i * g_step)
        b = int(start_b + i * b_step)
        color_hex = "{:02X}{:02X}{:02X}".format(r, g, b)
        gradient.append(color_hex)

    return gradient

# Generate the gradient from white to blue in 10 steps
gradient_white_to_blue = generate_gradient(start_color="#0000FF", end_color="#FFFFFF", num_steps=10)

# Generate the gradient from white to red in 10 steps
gradient_white_to_red = generate_gradient(start_color="#FF0000", end_color="#FFFFFF", num_steps=10)

# Create a sample DataFrame with integers between 1 and 100
data = {'Values': range(1, 101)}
df = pd.DataFrame(data)

# Create a Pandas Excel writer using openpyxl
writer = pd.ExcelWriter(outFile, engine='openpyxl')
df.to_excel(writer, sheet_name='Sheet1', index=False)

writer.close()


workbook = openpyxl.load_workbook(outFile)
worksheet = workbook['Sheet1']

worksheet['A1'].font = custom_font
# Set the zoom factor to 85% (0.85)
worksheet.sheet_view.zoomScale = 85

# Specify the top-left cell and block dimensions
start_cell = 'B1'
num_columns = 24
num_rows = 16

# Call the function to create the thick border
create_outer_thick_border(worksheet, start_cell, num_columns, num_rows)
create_plate_frame(worksheet, '1', start_cell, num_columns, num_rows)

start_cell = 'C26'
create_outer_thick_border(worksheet, start_cell, num_columns, num_rows)
create_plate_frame(worksheet, '2', start_cell, num_columns, num_rows)


# Iterate through the cells in the 'Values' column
for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=1)):
    cell = row[0]
    value = cell.value
    cell.font = custom_font

    # Apply the gradient fill based on the integer value
    if isinstance(value, int) and 1 <= value <= 100:
        
        percentage = (value - 1) / 99  # Normalize value to range [0, 1]
        if value < 11:
            cell.fill = PatternFill(start_color=gradient_white_to_red[value], end_color=gradient_white_to_red[value], fill_type="solid")
        elif value > 90:
            cell.fill = PatternFill(start_color=gradient_white_to_blue[100-value], end_color=gradient_white_to_blue[100-value], fill_type="solid")

    # Center-align the cell contents
    cell.alignment = Alignment(horizontal='center', vertical='center')


#workbook.close()
workbook.save(outFile)
