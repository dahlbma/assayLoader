import numpy as np
import pandas as pd
import numpy
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, PatternFill, Alignment, Border, Side, Font
from openpyxl.styles import Font
import io
from scipy.stats import percentileofscore

def addPlotToSheet(ws, cell, plt):
    plt.seek(0)
    img = Image(plt)
    img.width = 1400  # Set the width of the image in Excel
    img.height = 700  # Set the height of the image in Excel
    ws.add_image(img, cell)

    
def calculatePlateData(df, plate, ws):
    dataDf = df.loc[df['Type'].isna()]
    posDf = df.loc[df['Type'] == 'Pos']
    negDf = df.loc[df['Type'] == 'Neg']

    meanInhib = dataDf['Raw_data'].mean()
    stdInhib = dataDf['Raw_data'].std()

    meanPosCtrl = posDf['Raw_data'].mean()
    stdPosCtrl = posDf['Raw_data'].std()

    meanNegCtrl = negDf['Raw_data'].mean()
    stdNegCtrl = negDf['Raw_data'].std()

    pos3SD = 3 * stdPosCtrl
    neg3SD = 3 * stdNegCtrl

    Z = 1 - (pos3SD + neg3SD)/abs(meanPosCtrl - meanNegCtrl)
    
    condition = (df['Type'] == 'Neg') | (df['Type'] == 'Pos')
    df['inhibition'] = np.where(condition, None, 100*(1-(df['Raw_data']-meanPosCtrl)/(meanNegCtrl-meanPosCtrl)))

    df['negCtrlInhibition'] = np.where(df['Type'] == 'Neg', 100*(1-(df['Raw_data']-meanPosCtrl)/(meanNegCtrl-meanPosCtrl)), None)
    df['posCtrlInhibition'] = np.where(df['Type'] == 'Pos', 100*(1-(df['Raw_data']-meanPosCtrl)/(meanNegCtrl-meanPosCtrl)), None)

    return df, Z, meanPosCtrl, stdPosCtrl, meanNegCtrl, stdNegCtrl, meanInhib, stdInhib


def plotData(values, stds, sHeader):
    x_values = range(1, len(values) + 1 )
    plt.errorbar(x_values, values, yerr=stds, fmt='o', capsize=5,
                 label='Value with Std Dev')

    # Add labels and title
    plt.xlabel('Plate')
    plt.ylabel('Value')
    plt.title(sHeader)
    
    # Add a legend
    plt.legend()
    plt.grid()

    image_buffer = io.BytesIO()
    plt.savefig(image_buffer, format='png', dpi=300, bbox_inches='tight')
    plt.close()
    return image_buffer


def calcData(df, ws, heatMapWs):
    columns = ['Plate', 'meanRaw', 'stdRaw', 'meanNegCtrl', 'stdNegCtrl', 'meanPosCtrl', 'stdPosCtrl', 'Z-factor']
    df_summary = pd.DataFrame(columns=columns)
    df_inhibition = pd.DataFrame(columns=['inhibition'])
    df_inhibition_calculated = pd.DataFrame(columns=df.columns)

    listOfDfPlates = list()

    saPlates = list()
    for plate in df['plate'].unique():
        saPlates.append(plate)
        plate_df = df[df['plate'] == plate]
        (df_plate,
         Z,
         meanPosCtrl,
         stdPosCtrl,
         meanNegCtrl,
         stdNegCtrl,
         meanRaw,
         stdRaw) = calculatePlateData(plate_df, plate, ws)
        listOfDfPlates.append(df_plate)
        new_row = {'Plate': int(plate),
                   'meanRaw': meanRaw,
                   'stdRaw': stdRaw,
                   'meanNegCtrl': meanNegCtrl,
                   'stdNegCtrl': stdNegCtrl,
                   'meanPosCtrl': meanPosCtrl,
                   'stdPosCtrl': stdPosCtrl,
                   'Z-factor': Z}
        df_summary.loc[len(df_summary)] = new_row
        
        df_plate_inhibition = df_plate[(df_plate['Type'] != 'Neg') & (df_plate['Type'] != 'Pos')]
        df_inhibition = pd.concat([df_inhibition, df_plate_inhibition])
        df_inhibition_calculated = pd.concat([df_inhibition_calculated, df_plate])

    meanInhibition = df_inhibition['inhibition'].mean()
    stdInhibition = df_inhibition['inhibition'].std()
    hitLimit = meanInhibition + 3*stdInhibition
    minInhib = df_inhibition['inhibition'].min()
    maxInhib = df_inhibition['inhibition'].max()

    def calculate_hit(row):
        if row['inhibition'] is None:
            return 0
        elif row['inhibition'] > hitLimit:
            return 1
        else:
            return 0

    df_inhibition_calculated['hit'] = df_inhibition_calculated.apply(calculate_hit, axis=1)
    for row in dataframe_to_rows(df_inhibition_calculated, index=False, header=True):
        ws.append(row)

    ws['I1'] = f' Hit limit: {hitLimit}'
    ws['I2'] = f'Min inhib: {minInhib}'
    ws['I3'] = f'Max inhib: {maxInhib}'
    ws['I4'] = f'Mean inhib: {meanInhibition}'
    ws['I5'] = f'STD inhib: {stdInhibition}'
    '''
    print(minInhib)
    print(maxInhib)
    print(hitLimit)
    print(stdInhibition)
    print(meanInhibition)
    '''
    inhibPlt = plotData(df_summary['meanRaw'], df_summary['stdRaw'], 'Raw data')
    negPlt = plotData(df_summary['meanNegCtrl'], df_summary['stdNegCtrl'], 'NegCtrl')
    posPlt = plotData(df_summary['meanPosCtrl'], df_summary['stdPosCtrl'], 'PosCtrl')

    addPlotToSheet(ws, 'S1', inhibPlt)
    addPlotToSheet(ws, 'S40', negPlt)
    addPlotToSheet(ws, 'S80', posPlt)
    
    start_column = 'J'
    # Convert the DataFrame to rows and write them to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx + ord(start_column) - ord('A'))
            cell.value = value

    return listOfDfPlates, meanInhibition, stdInhibition

def create_plate_frame(ws, plate_id, top_left_cell, num_columns, num_rows):
    """
    Create an Excel sheet with plate_id, letters A-Z, and numbers 1-num_columns.

    Parameters:
    - ws: The Excel worksheet.
    - plate_id: The text to write in the top_left_cell.
    - top_left_cell: The top-left cell (e.g., 'A1').
    - num_columns: The number of columns for numbers (1-num_columns).
    - num_rows: The number of rows for letters (A-Z).
    """
    
    # Generate letters 'A' to 'Z' below the top_left_cell
    start_row, start_col = ws[top_left_cell].row, ws[top_left_cell].column
    
    # Write the plate_id to the specified cell
    ws[top_left_cell] = "Plate"
    cell = ws.cell(row=start_row, column=start_col+1)
    cell.value = plate_id

    for i in range(num_rows):
        letter = chr(ord('A') + i)
        cell = ws.cell(row=start_row + i + 3, column=start_col)
        cell.value = letter
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Generate numbers 1 to num_columns in a row, one column to the right
    for i in range(num_columns):
        number = i + 1
        cell = ws.cell(row=start_row + 2, column=start_col + i + 1)
        cell.value = number
        cell.alignment = Alignment(horizontal='center', vertical='center')


def create_outer_thick_border(ws, top_left_cell, num_columns, num_rows):
    """
    Create a thick border around the outer border of a block of cells in an Excel worksheet.

    Parameters:
    - ws: The Excel worksheet.
    - top_left_cell: The top-left cell (e.g., 'A1') of the block.
    - num_columns: The number of columns in the block.
    - num_rows: The number of rows in the block.
    """

    top_left_style = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style=None)
    )

    top_right_style = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style=None)
    )

    bottom_left_style = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thick')
    )

    bottom_right_style = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style='thick')
    )

    # Convert the top-left cell to row and column indices
    start_row, start_col = ws[top_left_cell].row + 3, ws[top_left_cell].column + 1

    top_l_cell = ws.cell(row=start_row, column=start_col)
    top_r_cell = ws.cell(row=start_row, column=start_col+num_columns-1)
    bottom_l_cell = ws.cell(row=start_row+num_rows-1, column=start_col)
    bottom_r_cell = ws.cell(row=start_row+num_rows-1, column=start_col+num_columns-1)

    # Apply thick border to top row
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=start_row, column=col)
        cell.border = Border(top=Side(style='thick'))
        top_left_cell = cell

    # Apply thick border to bottom row
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=start_row + num_rows - 1, column=col)
        cell.border = Border(bottom=Side(style='thick'))

    # Apply thick border to left column
    for row in range(start_row, start_row + num_rows):
        cell = ws.cell(row=row, column=start_col)
        cell.border = Border(left=Side(style='thick'))

    # Apply thick border to right column
    for row in range(start_row, start_row + num_rows):
        cell = ws.cell(row=row, column=start_col + num_columns - 1)
        cell.border = Border(right=Side(style='thick'))

    # Apply borders to the corner cells
    top_l_cell.border = top_left_style
    top_r_cell.border = top_right_style
    bottom_l_cell.border = bottom_left_style
    bottom_r_cell.border = bottom_right_style


def generate_gradient(end_color, start_color, num_steps):
    # Parse the start and end colors into RGB components
    start_r = int(start_color[1:3], 16)
    start_g = int(start_color[3:5], 16)
    start_b = int(start_color[5:7], 16)

    end_r = int(end_color[1:3], 16)
    end_g = int(end_color[3:5], 16)
    end_b = int(end_color[5:7], 16)

    # Calculate the step size for each RGB component
    r_step = (end_r - start_r) / num_steps
    g_step = (end_g - start_g) / num_steps
    b_step = (end_b - start_b) / num_steps

    # Generate the gradient colors and store them in a list
    gradient = []
    for i in range(num_steps + 1):
        r = int(start_r + i * r_step)
        g = int(start_g + i * g_step)
        b = int(start_b + i * b_step)
        color_hex = "40{:02X}{:02X}{:02X}".format(r, g, b)
        gradient.append(color_hex)

    return gradient


def populate_plate_data(heatMapsWs, plate, plateDf, start_cell):
    # Convert the top-left cell to row and column indices
    current_row , current_col = heatMapsWs[start_cell].row + 3, heatMapsWs[start_cell].column + 1
    plateDf['Raw_data'].fillna(0, inplace=True)
    plateDf['ptile'] = plateDf['Raw_data'].apply(lambda x: percentileofscore(plateDf['Raw_data'], x))
    
    for _, row in plateDf.iterrows():
        well = row['well']
        raw_data = row['Raw_data']

        try:
            percentile = int(row['ptile'])
        except Exception as e:
            percentile = 50
        
        # Split the well into row and column components (e.g., 'A01' -> 'A' and '01')
        well_row, well_col = well[0], int(well[1:])

        # Calculate the Excel row and column indices based on the well format
        excel_row = ord(well_row) - ord('A') + 1
        excel_col = well_col

        # Insert the Raw_data value into the corresponding cell
        #heatMapsWs.cell(row=excel_row + current_row - 1, column=excel_col + current_col - 1, value=raw_data)
        cell = heatMapsWs.cell(row=current_row, column=current_col , value=raw_data)
        cell.fill = PatternFill(start_color=color_list[percentile], end_color=color_list[percentile], fill_type="solid")

        #print(well, current_row, current_col, raw_data)
        # Move to the next column
        current_col += 1

        # If we've reached the 24th column, move to the next row and reset the column counter
        if current_col > 25:
            current_row += 1
            current_col = 2
    #quit()


pd.set_option('mode.chained_assignment', None)

decimal_style = NamedStyle(name='two_decimals')
decimal_style.number_format = '0.00'
# Define a custom font (Calibri 10pt)
custom_font = Font(name='Calibri', size=10)

excel_file_path = 'screenresults.xlsx'
df = pd.read_csv("plate_Raw_data.csv", delimiter='\t')



# Generate the gradient from white to red in 10 steps
gradient_white_to_red = generate_gradient(start_color="#FF0000", end_color="#FFFFFF", num_steps=20)

# Generate the gradient from white to blue in 10 steps
gradient_white_to_blue = generate_gradient(start_color="#0000FF", end_color="#FFFFFF", num_steps=20)

white_list = ['FFFFFF'] * 60
color_list = gradient_white_to_red + white_list + gradient_white_to_blue

wb = Workbook()
screenDataWs = wb.active
screenDataWs.title = 'ScreenDataAnalysis'
# Set the zoom factor to 85% (0.85)
screenDataWs.sheet_view.zoomScale = 85


#########################################################
# Add a new sheet named "Heat maps"
heatMapsWs = 'Heat maps'
heatMapsWs = wb.create_sheet(title=heatMapsWs)
heatMapsWs.sheet_view.zoomScale = 85

start_col = 'A'
start_row = 1
num_columns = 24
num_rows = 16

iPlateRows = 21
# End Heat map data
#########################################################

listOfPlatesDf, meanInhibition, stdInhibition = calcData(df, screenDataWs, heatMapsWs)

for column in screenDataWs.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Get the column letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = max(max_length + 2, 10)  # Minimum width of 10 characters
    screenDataWs.column_dimensions[column_letter].width = adjusted_width

for row in screenDataWs.iter_rows():
    for cell in row:
        cell.style = decimal_style
        cell.font = custom_font


for plateDf in listOfPlatesDf:
    # Call the function to create the thick border
    plate = plateDf.iloc[0]['plate']

    iRow = start_row + ((plate-1) * iPlateRows)
    start_cell = start_col + str(iRow)
    create_outer_thick_border(heatMapsWs, start_cell, num_columns, num_rows)
    create_plate_frame(heatMapsWs, plate, start_cell, num_columns, num_rows)
    populate_plate_data(heatMapsWs, plate, plateDf, start_cell)











wb.save(excel_file_path)








'''
100*(1-(A31-J216)/(G24-J216)))

Percent Inhibition:
100 * (1- (DataPoint - AVG(PosCtrl))/(AVG(NegCtrl)-AVG(PosCtrl)))

NegCtrl, 0% Control (F24):
100 * (1-(NegCtrlDataPoint-AVG(PosCtrl)) /(Avg(NegCtrl) - AVG(PosCtrl)))


PosCtrl, 100% Ctrl, 100*(1-(A25-J216)/(G24-J216)))
I25:


StdDev(NegCtrl): H24
G24 = Avg(NegCtrl)
J216 = AVG(PosCtrl)
K216 = STD(PosCtrl)

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Heatmap"


# Set the background color of cell A1 to red
cell = sheet['A1']
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
cell.fill = red_fill

# Save the workbook to a file
workbook.save('screenresults.xlsx')


'''
