import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, PatternFill, Alignment, Border, Side, Font
import io
from scipy.stats import percentileofscore
from PyQt5.QtWidgets import QApplication

# 1536-well plate (32 rows and 48 columns) would span from well "A01" to well "AF48".

def addPlotToSheet(ws, cell, plt):
    plt.seek(0)
    img = Image(plt)
    img.width = 1400  # Set the width of the image in Excel
    img.height = 700  # Set the height of the image in Excel
    ws.add_image(img, cell)


def addLineOfDataToSheet(ws, sText, start_cell, df_data, data_column):
    # Insert sText into the start_cell
    ws[start_cell] = sText

    # Get the row number from the start_cell
    row_number, col_number = coordinate_to_tuple(start_cell)
    # Convert the DataFrame to a list of rows

    data_list = df_data[data_column].tolist()
    for point in data_list:
        col_number += 1
        ws.cell(row=row_number, column=col_number, value=point)


def addColumnOfDataToSheet(ws, sText, start_cell, df_data, data_column):
    # Insert sText into the start_cell
    ws[start_cell] = sText

    # Get the row number from the start_cell
    row_number, col_number = coordinate_to_tuple(start_cell)
    # Convert the DataFrame to a list of rows

    data_list = df_data[data_column].tolist()
    for point in data_list:
        row_number += 1
        ws.cell(row=row_number, column=col_number, value=point)


def setBackgroundColor(ws, color, start_cell, end_cell):
    """
    Set the background color of a range of cells in an Excel worksheet.

    Parameters:
    - ws: Excel worksheet (openpyxl worksheet object)
    - color: Hex value of the background color (e.g., '#FF5050')
    - start_cell: Upper left corner of the block to color (e.g., 'A1')
    - end_cell: Lower right corner of the block to color (e.g., 'B4')
    """

    # Create a PatternFill object with the specified color
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Get the column and row indices of the start and end cells
    start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
    end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)

    # Loop through the range of cells and set the background color
    for row in ws.iter_rows(min_row=start_row,
                            max_row=end_row,
                            min_col=openpyxl.utils.column_index_from_string(start_col),
                            max_col=openpyxl.utils.column_index_from_string(end_col)):
        for cell in row:
            cell.fill = fill

def calculatePlateData(df, plate, ws):
    dataDf = df.loc[df['type'] == 'Data']
    posDf = df.loc[df['type'] == 'Pos']
    negDf = df.loc[df['type'] == 'Neg']

    meanInhib = dataDf['raw_data'].mean()
    stdInhib = dataDf['raw_data'].std()

    meanPosCtrl = posDf['raw_data'].mean()
    stdPosCtrl = posDf['raw_data'].std()

    meanNegCtrl = negDf['raw_data'].mean()
    stdNegCtrl = negDf['raw_data'].std()

    pos3SD = 3 * stdPosCtrl
    neg3SD = 3 * stdNegCtrl
    try:
        Z = 1 - (pos3SD + neg3SD)/abs(meanPosCtrl - meanNegCtrl)
    except:
        printQcLog(f'Failed calculating plate: {plate}')
        return False
    condition = (df['type'] == 'Neg') | (df['type'] == 'Pos')
    df['inhibition'] = np.where(condition, None, 100*(1-(df['raw_data']-meanPosCtrl)/(meanNegCtrl-meanPosCtrl)))

    df['negCtrlInhibition'] = np.where(df['type'] == 'Neg', 100*(1-(df['raw_data']-meanPosCtrl)/(meanNegCtrl-meanPosCtrl)), None)
    df['posCtrlInhibition'] = np.where(df['type'] == 'Pos', 100*(1-(df['raw_data']-meanPosCtrl)/(meanNegCtrl-meanPosCtrl)), None)
    
    return df, Z, meanPosCtrl, stdPosCtrl, meanNegCtrl, stdNegCtrl, meanInhib, stdInhib

def plotZfactor(df):
    # Create a bar plot for the 'Z-factor' column
    plt.bar(range(len(df['Plate'])), df['Z-factor'])
    
    # Set labels and title
    plt.xlabel('Plate')
    plt.ylabel('Z-factor')
    plt.title('Bar Plot of Z-factor')

    plt.xticks(fontsize=6)
    plt.xticks(range(1, len(df) +1, 3), fontsize=6)
    
    image_buffer = io.BytesIO()
    plt.savefig(image_buffer, format='png', dpi=300, bbox_inches='tight')
    plt.close()
    return image_buffer


def inhibitionScatterPlot(df_inhibition, hitLimit):
    # Create a scatterplot of the "inhibition" columns
    plt.scatter(range(len(df_inhibition)), df_inhibition['posCtrlInhibition'], label='PosCtrl', marker='.', s=1, c='green')
    plt.scatter(range(len(df_inhibition)), df_inhibition['inhibition'], label='Inhibition', marker='.', s=2, c='blue')
    plt.scatter(range(len(df_inhibition)), df_inhibition['negCtrlInhibition'], label='NegCtrl', marker='.', s=1, c='red')
    # Draw a horizontal line at the hit limit
    plt.axhline(y=hitLimit, color='red', linestyle='--', label='Hit Limit')

    # Set labels and title
    plt.xlabel('Data Points')
    plt.ylabel('Inhibition')
    plt.title('Inhibition scatterplot')
    
    # Add a legend
    plt.legend()
    
    image_buffer = io.BytesIO()
    plt.savefig(image_buffer, format='png', dpi=300, bbox_inches='tight')
    plt.close()
    return image_buffer


def plotInhibitionHistogram(df_inhibition):
    max_value = int(df_inhibition['inhibition'].max())
    min_value = int(df_inhibition['inhibition'].min())
    
    plt.hist(df_inhibition['inhibition'], bins=max_value - min_value, edgecolor='black')
    plt.xlabel('Inhibition')
    plt.ylabel('Frequency')
    plt.title('Distribution of Inhibition')

    image_buffer = io.BytesIO()
    plt.savefig(image_buffer, format='png', dpi=300, bbox_inches='tight')
    plt.close()
    return image_buffer


def plotMeanStd(values, stds, sHeader):
    x_values = range(1, len(values) + 1 )
    plt.errorbar(x_values, values, yerr=stds, fmt='o', capsize=5,
                 label='Mean value with Std Dev')

    # Add labels and title
    plt.xlabel('Plate')
    plt.ylabel('Mean')
    plt.title(sHeader)
    
    # Add a legend
    plt.legend()
    plt.grid()

    image_buffer = io.BytesIO()
    plt.savefig(image_buffer, format='png', dpi=300, bbox_inches='tight')
    plt.close()
    return image_buffer


def calcData(self, excelSettings, df, ws, heatMapWs, iHitThreshold):
    columns = ['Plate', 'meanRaw', 'stdRaw', 'meanNegCtrl', 'stdNegCtrl', 'meanPosCtrl', 'stdPosCtrl', 'Z-factor']
    df_summary = pd.DataFrame(columns=columns)
    df_inhibition = pd.DataFrame(columns=['inhibition'])
    df_inhibition_calculated = pd.DataFrame(columns=df.columns)

    listOfDfPlates = list()
    
    saPlates = list()
    for plate in df['plate'].unique():
        self.printQcLog(f"Reading plate {plate}")

        saPlates.append(plate)
        plate_df = df[df['plate'] == plate]
        (df_plate,
         Z,
         meanPosCtrl,
         stdPosCtrl,
         meanNegCtrl,
         stdNegCtrl,
         meanRaw,
         stdRaw) = calculatePlateData(plate_df, plate, ws)
        listOfDfPlates.append(df_plate)
        new_row = {'Plate': plate,
                   'meanRaw': meanRaw,
                   'stdRaw': stdRaw,
                   'meanNegCtrl': meanNegCtrl,
                   'stdNegCtrl': stdNegCtrl,
                   'meanPosCtrl': meanPosCtrl,
                   'stdPosCtrl': stdPosCtrl,
                   'Z-factor': Z}
        df_summary.loc[len(df_summary)] = new_row
        
        df_plate_inhibition = df_plate[(df_plate['type'] != 'Neg') & (df_plate['type'] != 'Pos')]
        df_inhibition = pd.concat([df_inhibition, df_plate_inhibition])
        df_inhibition_calculated = pd.concat([df_inhibition_calculated, df_plate])

    meanInhibition = df_inhibition['inhibition'].mean()
    stdInhibition = df_inhibition['inhibition'].std()
    if iHitThreshold == -1000:
        hitLimit = meanInhibition + 3*stdInhibition
    else:
        hitLimit = float(iHitThreshold)
    minInhib = df_inhibition['inhibition'].min()
    maxInhib = df_inhibition['inhibition'].max()

    def calculate_hit(row):
        if row['inhibition'] is None:
            return 0
        elif row['inhibition'] > hitLimit:
            return 1
        else:
            return 0

    df_inhibition_calculated['hit'] = df_inhibition_calculated.apply(calculate_hit, axis=1)
    for row in dataframe_to_rows(df_inhibition_calculated, index=False, header=True):
        ws.append(row)
    
    ws['K1'] = ' Hit limit: {:.2f}'.format(hitLimit)
    ws['K2'] = ' Min inhib: {:.2f}'.format(minInhib)
    ws['K3'] = ' Max inhib: {:.2f}'.format(maxInhib)
    ws['K4'] = ' Mean inhib: {:.2f}'.format(meanInhibition)
    ws['K5'] = ' STD inhib: {:.2f}'.format(stdInhibition)

    inhibPlt = plotMeanStd(df_summary['meanRaw'], df_summary['stdRaw'], 'Raw data')
    negPlt = plotMeanStd(df_summary['meanNegCtrl'], df_summary['stdNegCtrl'], 'NegCtrl')
    posPlt = plotMeanStd(df_summary['meanPosCtrl'], df_summary['stdPosCtrl'], 'PosCtrl')
    zFactorPlt = plotZfactor(df_summary)
    inhibitionHistogramPlt = plotInhibitionHistogram(df_inhibition)
    inhibitionScatterPlt = inhibitionScatterPlot(df_inhibition_calculated, hitLimit)
    
    addPlotToSheet(ws, 'U1', inhibPlt)
    addPlotToSheet(ws, 'U40', negPlt)
    addPlotToSheet(ws, 'U80', posPlt)
    addPlotToSheet(ws, 'U120', inhibitionHistogramPlt)
    addPlotToSheet(ws, 'U160', inhibitionScatterPlt)
    addPlotToSheet(ws, 'U200', zFactorPlt)
    
    start_column = 'L'
    #light_red_3_fill = PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")
    light_red_3_fill = PatternFill(start_color="11FF5050", end_color="11FF5050", fill_type="solid")
    # Convert the DataFrame to rows and write them to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx + ord(start_column) - ord('A'))
            cell.value = value
            cell.font = excelSettings["custom_font"]
            cell.style = excelSettings["decimal_style"]

    return listOfDfPlates, meanInhibition, stdInhibition, df_inhibition_calculated, hitLimit


def create_plate_frame(ws, sHeading, plate_id, top_left_cell, num_columns, num_rows):
    """
    Create an Excel sheet with plate_id, letters A-Z, and numbers 1-num_columns.

    Parameters:
    - ws: The Excel worksheet.
    - plate_id: The text to write in the top_left_cell.
    - top_left_cell: The top-left cell (e.g., 'A1').
    - num_columns: The number of columns for numbers (1-num_columns).
    - num_rows: The number of rows for letters (A-Z).
    """
    
    # Generate letters 'A' to 'Z' below the top_left_cell
    start_row, start_col = ws[top_left_cell].row, ws[top_left_cell].column
    
    # Write the plate_id to the specified cell
    ws[top_left_cell] = sHeading
    cell = ws.cell(row=start_row, column=start_col+1)
    cell.value = plate_id

    for i in range(num_rows):
        letter = chr(ord('A') + i)
        cell = ws.cell(row=start_row + i + 3, column=start_col)
        cell.value = letter
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Generate numbers 1 to num_columns in a row, one column to the right
    for i in range(num_columns):
        number = i + 1
        cell = ws.cell(row=start_row + 2, column=start_col + i + 1)
        cell.value = number
        cell.alignment = Alignment(horizontal='center', vertical='center')


def create_outer_thick_border(ws, top_left_cell, num_columns, num_rows):
    """
    Create a thick border around the outer border of a block of cells in an Excel worksheet.

    Parameters:
    - ws: The Excel worksheet.
    - top_left_cell: The top-left cell (e.g., 'A1') of the block.
    - num_columns: The number of columns in the block.
    - num_rows: The number of rows in the block.
    """

    top_left_style = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style='thick'),
        bottom=Side(style=None)
    )

    top_right_style = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style=None)
    )

    bottom_left_style = Border(
        left=Side(style='thick'),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style='thick')
    )

    bottom_right_style = Border(
        left=Side(style=None),
        right=Side(style='thick'),
        top=Side(style=None),
        bottom=Side(style='thick')
    )

    # Convert the top-left cell to row and column indices
    start_row, start_col = ws[top_left_cell].row + 3, ws[top_left_cell].column + 1

    top_l_cell = ws.cell(row=start_row, column=start_col)
    top_r_cell = ws.cell(row=start_row, column=start_col+num_columns-1)
    bottom_l_cell = ws.cell(row=start_row+num_rows-1, column=start_col)
    bottom_r_cell = ws.cell(row=start_row+num_rows-1, column=start_col+num_columns-1)

    # Apply thick border to top row
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=start_row, column=col)
        cell.border = Border(top=Side(style='thick'))
        top_left_cell = cell

    # Apply thick border to bottom row
    for col in range(start_col, start_col + num_columns):
        cell = ws.cell(row=start_row + num_rows - 1, column=col)
        cell.border = Border(bottom=Side(style='thick'))

    # Apply thick border to left column
    for row in range(start_row, start_row + num_rows):
        cell = ws.cell(row=row, column=start_col)
        cell.border = Border(left=Side(style='thick'))

    # Apply thick border to right column
    for row in range(start_row, start_row + num_rows):
        cell = ws.cell(row=row, column=start_col + num_columns - 1)
        cell.border = Border(right=Side(style='thick'))

    # Apply borders to the corner cells
    top_l_cell.border = top_left_style
    top_r_cell.border = top_right_style
    bottom_l_cell.border = bottom_left_style
    bottom_r_cell.border = bottom_right_style


def generate_gradient(end_color, start_color, num_steps):
    # Parse the start and end colors into RGB components
    start_r = int(start_color[1:3], 16)
    start_g = int(start_color[3:5], 16)
    start_b = int(start_color[5:7], 16)

    end_r = int(end_color[1:3], 16)
    end_g = int(end_color[3:5], 16)
    end_b = int(end_color[5:7], 16)

    # Calculate the step size for each RGB component
    r_step = (end_r - start_r) / num_steps
    g_step = (end_g - start_g) / num_steps
    b_step = (end_b - start_b) / num_steps

    # Generate the gradient colors and store them in a list
    gradient = []
    for i in range(num_steps + 1):
        r = int(start_r + i * r_step)
        g = int(start_g + i * g_step)
        b = int(start_b + i * b_step)
        color_hex = "{:02X}{:02X}{:02X}".format(r, g, b)
        gradient.append(color_hex)

    gradient.reverse()
    return gradient


def populate_plate_data(excelSettings, heatMapsWs, plate, plateDf, start_cell, data_col, lDebug=False):
    whiteFont = Font(name='Calibri', color="FFFFFF")
    # Convert the top-left cell to row and column indices
    current_row , current_col = heatMapsWs[start_cell].row + 3, heatMapsWs[start_cell].column + 1
    start_col = current_col
    start_row = current_row
    
    # Extract the numeric part as "col" and the letter part as "row"
    plateDf['col'] = plateDf['well'].str.extract('(\d+)').astype(int)
    plateDf['row'] = plateDf['well'].str.extract('([A-Z])').iloc[:, 0].apply(lambda x: ord(x[0]) - ord('A') + 1)

    if(lDebug):
        pass

    plateDf = plateDf.dropna(subset=[data_col])

    # Define a function to calculate the percentile, handling null values
    def calculate_percentile(x):
        if x is not None:
            return percentileofscore(plateDf[data_col].dropna(), x)
        else:
            return None

    # Apply the function to the 'data_col' and create the 'ptile' column
    plateDf['ptile'] = plateDf[data_col].apply(calculate_percentile)
    try:
        plateDf['ptile'] = plateDf['ptile'].astype(int)
    except Exception as e:
        # Catch the exception and print the error message
        self.printQcLog(f"An error occurred: {e}", 'error')

    for _, row in plateDf.iterrows():
        well = row['well']
        raw_data = row[data_col]
        percentile = min(row['ptile'], 99)

        # Split the well into row and column components (e.g., 'A01' -> 'A' and '01')
        well_row, well_col = well[0], int(well[1:])

        # Calculate the Excel row and column indices based on the well format
        excel_row = ord(well_row) - ord('A') + 1
        excel_col = well_col
        
        # Insert the raw_data value into the corresponding cell
        cell = heatMapsWs.cell(row=start_row + row['row'] -1, column=start_col + row['col'] -1, value=raw_data)
        cell.fill = PatternFill(start_color=excelSettings["color_list"][percentile],
                                end_color=excelSettings["color_list"][percentile],
                                fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if percentile < 10 or percentile > 90:
            cell.font = whiteFont

def calcQc(self, input_file, output_file, iHitThreshold):
        
    pd.set_option('mode.chained_assignment', None)
    df = pd.read_csv(input_file, delimiter='\t')
    self.printQcLog(f"Reading input")
    
    # Generate the gradient from white to red in 10 steps
    #gradient_white_to_red = generate_gradient(start_color="#FFFFFF", end_color="#AF0000", num_steps=24)
    gradient_white_to_red = generate_gradient(start_color="#FFFFFF", end_color="#C0504D", num_steps=34)

    # Generate the gradient from white to blue in 10 steps
    #gradient_white_to_blue = generate_gradient(start_color="#0000AF", end_color="#FFFFFF", num_steps=24)
    gradient_white_to_blue = generate_gradient(start_color="#1F497D", end_color="#FFFFFF", num_steps=34)
    white_list = ['FFFFFF'] * 30

    excel_file_path = output_file
    wb = Workbook()
    screenDataWs = wb.active
    screenDataWs.title = 'ScreenDataAnalysis'

    # Set the zoom factor to 80% (0.80)
    screenDataWs.sheet_view.zoomScale = 80

    excelSettings = {
        "decimal_style": NamedStyle(name='two_decimals'),
        "decimal_style.number_format": '0.00',
        "custom_font": Font(name='Calibri', size=10),
        "color_list": gradient_white_to_red + white_list + gradient_white_to_blue
    }

    #########################################################
    # Add a new sheet named "Heat maps"
    heatMapsWs = 'Heat maps'
    heatMapsWs = wb.create_sheet(title=heatMapsWs)
    heatMapsWs.sheet_view.zoomScale = 80

    start_col = 'A'
    start_row = 1
    num_columns = 24
    num_rows = 16

    iPlateRows = 21
    # End Heat map data
    #########################################################

    self.printQcLog(f"Calculating all means and std")
    listOfPlatesDf, meanInhibition, stdInhibition, dfCalcData, hitLimt= calcData(self,
                                                                                 excelSettings,
                                                                                 df,
                                                                                 screenDataWs,
                                                                                 heatMapsWs,
                                                                                 iHitThreshold)
    self.printQcLog(f"All data read")

    for column in screenDataWs.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(max_length + 2, 10)  # Minimum width of 10 characters
        screenDataWs.column_dimensions[column_letter].width = adjusted_width

    for row in screenDataWs.iter_rows():
        for cell in row:
            cell.style = excelSettings["decimal_style"]
            cell.font = excelSettings["custom_font"]

    iPlate = 0
    for plateDf in listOfPlatesDf:
        # Call the function to create the thick border
        plate = plateDf.iloc[0]['plate']
        self.printQcLog(f"Calculating stats for {plate}")

        iRow = start_row + ((iPlate) * iPlateRows)
    
        start_cell = start_col + str(iRow)
        create_outer_thick_border(heatMapsWs, start_cell, num_columns, num_rows)
        create_plate_frame(heatMapsWs, 'Plate', plate, start_cell, num_columns, num_rows)
        populate_plate_data(excelSettings, heatMapsWs, plate, plateDf, start_cell, 'raw_data')
        iPlate += 1

    # Group by 'well' and sum the 'hit' column to count occurrences of '1'
    df_hit_distr = dfCalcData.groupby('well')['hit'].sum().reset_index()
    # Rename the 'hit' column to 'count'
    df_hit_distr = df_hit_distr.rename(columns={'hit': 'count'})

    ######################################################
    ##  Hit count per well
    start_cell = 'N240'
    create_outer_thick_border(screenDataWs, start_cell, num_columns, num_rows)
    create_plate_frame(screenDataWs, 'Hit Distr', "", start_cell, num_columns, num_rows)
    populate_plate_data(excelSettings, screenDataWs, 1, df_hit_distr, start_cell, 'count')
    ##
    ######################################################


    ######################################################
    ##  Average raw_data value for each well
    df_avg_well = df.groupby("well")["raw_data"].mean().reset_index()
    df_avg_well.rename(columns={"raw_data": "avgValue"}, inplace=True)


    new_data = {
        "well": [f"{row}{col:02d}" for row in "ABCDEFGHIJKLMNOP" for col in range(1, 25)]
    }

    for type_value in ["Data", "Neg", "Pos"]:
        avg_values = df[df["type"] == type_value].groupby("well")["raw_data"].mean()    
        new_data[f"avg{type_value}Value"] = [avg_values.get(well, None) for well in new_data["well"]]

    df_avg_well = pd.DataFrame(new_data)

    # Ensure that the new DataFrame has all 384 well values (A01-P24)
    all_wells = [f"{row}{col:02d}" for row in "ABCDEFGHIJKLMNOP" for col in range(1, 25)]
    df_avg_well = df_avg_well.reindex(columns=["well", "avgDataValue"])
    df_avg_well["well"] = all_wells

    start_cell = 'N266'
    create_outer_thick_border(screenDataWs, start_cell, num_columns, num_rows)
    create_plate_frame(screenDataWs, 'Well Avg', "", start_cell, num_columns, num_rows)
    populate_plate_data(excelSettings, screenDataWs, 1, df_avg_well, start_cell, 'avgDataValue', lDebug=True)

    self.printQcLog(f"Generate Excel file")

    ##
    ######################################################

    setBackgroundColor(ws=screenDataWs, color="32CD32", start_cell='K1', end_cell='K5')
    setBackgroundColor(ws=screenDataWs, color="ffd7d7", start_cell='L1', end_cell='S' + str(len(listOfPlatesDf) + 1))

    wb.save(excel_file_path)

    return hitLimt, dfCalcData
