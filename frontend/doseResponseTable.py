from scipy.optimize import curve_fit
from scipy.integrate import quad
import os
import math
import numpy as np
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QApplication, QFileDialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from scatterplotWidget import ScatterplotWidget
import assaylib
import configParams as cfg
import warnings
warnings.filterwarnings('ignore')

GRAPH_WIDTH = cfg.GRAPH_WIDTH
GRAPH_HEIGHT = cfg.GRAPH_HEIGHT

class DoseResponseTable(QTableWidget):
    def __init__(self, inputWidget):
        # What is the inputWidget, it has type QWidget ??????
        super(DoseResponseTable, self).__init__()
        self.ic50_col = 2
        self.ic50std_col = 3
        self.slope_col = 4
        self.bottom_col = 5
        self.top_col = 6
        self.minConc_col = 7
        self.maxConc_col = 8
        self.auc_col = 9
        self.icmax_col = 10
        self.graph_col = 11
        self.confirmed_col = 12
        self.comment_col = 13
        self.workingDirectory = ''
        self.parent = None

        # Enable saveExcel_btn when any cell is edited
        self.cellChanged.connect(self._enable_save_excel_btn)


    def _enable_save_excel_btn(self, row, column):
        if self.parent.saveExcel_btn:
            self.parent.saveExcel_btn.setEnabled(True)

    def reset_table_data(self):
        """
        Resets the table by clearing all existing data and widgets,
        then optionally populates it with new_df.
        Ensures proper memory cleanup for embedded widgets.
        """
        
        # 1. Clear existing ScatterplotWidgets and ensure they are deleted
        num_rows = self.rowCount()
        for row in range(num_rows):
            widget_item = self.cellWidget(row, self.graph_col)
            if isinstance(widget_item, ScatterplotWidget):
                # Remove the widget from the cell
                self.removeCellWidget(row, self.graph_col)
                # Schedule the widget for deletion. This is CRUCIAL for memory cleanup.
                widget_item.deleteLater()
                # print(f"  - Scheduled ScatterplotWidget in row {row} for deletion.")

        # 2. Clear all table contents and reset row count
        self.clearContents() # Clears text in cells
        self.setRowCount(0)   # Resets the number of rows

        print("Existing table data and widgets cleared.")

    
    def qtablewidget_to_dataframe(self):
        """
        Converts a QTableWidget to a pandas DataFrame.
        
        Args:
        table_widget (QTableWidget): The QTableWidget to convert.
        
        Returns:
        pd.DataFrame: A pandas DataFrame containing the QTableWidget data.
        """
        
        data = []
        headers = []

        # Get headers
        for column in range(self.columnCount()):
            headers.append(self.horizontalHeaderItem(column).text())
        headers.append('comment')

        # Get data
        for row in range(self.rowCount()):

            row_data = []
            sComment = ''
            for column in range(self.columnCount()):
                item = self.item(row, column)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append(None)  # Handle empty cells
                
            row_data.append(sComment)
            data.append(row_data)
            
        df = pd.DataFrame(data, columns=headers)
        #df = self.generateComment(df)

        '''
        mask = ~df['Compound'].str.startswith('CBK')
        # Remove the rows that doesn't start with 'CBK'
        df = df[~mask]
        '''
        return df

        
    def generate_scatterplots(self, file_path, yScale, parent):
        self.parent = parent
        outputDir = os.path.dirname(file_path)
        self.workingDirectory = outputDir
        self.reset_table_data()
        
        self.setColumnWidth(self.graph_col, GRAPH_WIDTH)
        df = pd.read_excel(file_path)
        df = df[df['Compound ID'].str.startswith('CBK')] # Remove all controls

        dialog = assaylib.CancelDialog(self)
        dialog.show()

        iTotPlotNum = len(df.groupby('Batch nr'))
        iCount = 0
        for batch_nr, batch_df in df.groupby('Batch nr'):
            iCount += 1
            dialog.update_label(f"{iCount + 1} of {iTotPlotNum} curves")
            if dialog.cancelled:
                break
            
            rowPosition = self.rowCount()
            if parent.test == 'true':
                if rowPosition > 20:
                    continue
            self.insertRow(rowPosition)
            self.setRowHeight(rowPosition, GRAPH_HEIGHT)
            
            print(f'Plot nr: {rowPosition}')
            # Call the scatter function for each batch
            self.plotCurve(batch_df, rowPosition, yScale)

        total_param_changes = 0
        for row in range(self.rowCount()):
            widget = self.cellWidget(row, self.graph_col)
            if widget and hasattr(widget, 'iParameterCount'):
                total_param_changes += widget.iParameterCount
        print("Total parameter changes:", total_param_changes)
        dialog.close()
        self.saveToExcel('DR_Excel.xlsx')
        return batch_df


    def changeIC50_EC50_heading(self, newHeading):
        header_item = self.horizontalHeaderItem(self.ic50_col)
        assaylib.printDbg(newHeading)
        if header_item:
            header_item.setText(newHeading)


    def plotCurve(self, batch_df, rowPosition, yScale):
        # "Batch nr" "Compound ID" "finalConc_nM" "yMean" "yStd"
        batch = batch_df['Batch nr'].iloc[0]
        compound = batch_df['Compound ID'].iloc[0]

        scatterplot_widget = ScatterplotWidget(batch_df, rowPosition, yScale, self.workingDirectory)
        scatterplot_widget.fit_curve_to_data()
        scatterplot_widget.plot_curve()
        item = QTableWidgetItem(batch)
        self.setItem(rowPosition, 0, item)

        item = QTableWidgetItem(compound)
        self.setItem(rowPosition, 1, item)

        item = QTableWidgetItem()
        self.setItem(rowPosition, self.graph_col, item)
        self.setCellWidget(rowPosition, self.graph_col, scatterplot_widget)
        self.setCurrentCell(rowPosition, self.graph_col)
        self.updateTable(rowPosition, scatterplot_widget)

        QApplication.processEvents()

    def updateMaxConc(self, row, maxConc):
        item = QTableWidgetItem(str(f"{maxConc:.1f}"))
        self.setItem(row, self.maxConc_col, item)
        ### Mats
        ### On next line we need to find the inhibition for the higest concentration when we have deselected concs.
        # self.setItem(row, self.icmax_col, item)


    def updateTable(self, rowPosition, scatterplot_widget):
        item = QTableWidgetItem(str("{:.2e}".format(scatterplot_widget.ic50)))
        self.setItem(rowPosition, self.ic50_col, item) # 2

        item = QTableWidgetItem(scatterplot_widget.sInfo)
        self.setItem(rowPosition, self.ic50std_col, item) # 3

        item = QTableWidgetItem(str(f"{scatterplot_widget.slope:.2f}"))
        self.setItem(rowPosition, self.slope_col, item) #4

        item = QTableWidgetItem(str(f"{scatterplot_widget.bottom:.2f}"))
        self.setItem(rowPosition, self.bottom_col, item) # 5

        item = QTableWidgetItem(str(f"{scatterplot_widget.top:.2f}"))
        self.setItem(rowPosition, self.top_col, item)

        item = QTableWidgetItem(str(f"{scatterplot_widget.minConc:.1f}"))
        self.setItem(rowPosition, self.minConc_col, item)

        item = QTableWidgetItem(str(f"{scatterplot_widget.maxConc:.1f}"))
        self.setItem(rowPosition, self.maxConc_col, item)

        item = QTableWidgetItem(str(f"{scatterplot_widget.icmax:.2f}"))
        self.setItem(rowPosition, self.icmax_col, item)

        item = QTableWidgetItem(str(f"{scatterplot_widget.auc:.5f}"))
        self.setItem(rowPosition, self.auc_col, item)

        item = QTableWidgetItem(str(f"{scatterplot_widget.confirmed}"))
        self.setItem(rowPosition, self.confirmed_col, item)

        item = QTableWidgetItem(str(f"{scatterplot_widget.comment}"))
        self.setItem(rowPosition, self.comment_col, item)


    def saveToExcel(self, sFileName = None):
        sDir = self.workingDirectory
        sFile = 'DR_Excel.xlsx'
        file_path = os.path.join(sDir, sFile)

        imgDir = self.workingDirectory + '/img'
        if not os.path.exists(imgDir):
            # Create the directory and any missing parent directories
            os.makedirs(imgDir)
        
        if sFileName in (None, False):
            sDefaultFile = os.path.join(sDir, sFile)
            # Open a file dialog to select the location and file name for saving
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(self, "Save As", sDefaultFile,
                                                       "Excel Files (*.xlsx *.xls)",
                                                       options=options)

            file_path = file_name
            if file_name in ('', None):
                return
            sDir = os.path.dirname(file_name)

            #imgDir = sDir + '/img'
            if not os.path.exists(imgDir):
                # Create the directory and any missing parent directories
                os.makedirs(imgDir)

        wb = Workbook()
        ws = wb.active

        headings = ["Batch", "Compound", "IC50", "Fit quality", "Slope",
                    "Bottom", "Top", "MinConc nM", "MaxConc nM", "AUC", "ICMax", "Graph", "Confirmed", "Comment"]
                
        # Convert QTableWidget data to a pandas DataFrame
        table_data = []
        for row in range(self.rowCount()):
            row_data = [self.item(row, col).text() if col < len(headings) else None for col in range(self.columnCount())]
            table_data.append(row_data)


        df = pd.DataFrame(table_data, columns=headings)

        for col_num, heading in enumerate(headings, 1):
            cell = ws.cell(row=1, column=col_num, value=heading)
            cell.font = Font(bold=True)

        # Write DataFrame to Excel
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start from row 2 to leave space for header
            for c_idx, value in enumerate(row, start=1):
                if c_idx > 2:
                    try:
                        if c_idx in (3, 4):
                            ws.cell(row=r_idx, column=c_idx, value=float(f'{value:e}'))
                        else:
                            ws.cell(row=r_idx, column=c_idx, value=float(value))
                    except:
                        ws.cell(row=r_idx, column=c_idx, value=value)
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Add scatterplots to Excel
        for i, canvas in enumerate(df['Graph']):
            image_path = f"{imgDir}/{i}.png"
            img = Image(image_path)
                        
            #ws.add_image(img)
            ws.add_image(img, f"L{i + 2}")

        # Adjust row heights and column widths to fit images
        for r_idx in range(2, len(df) + 2):  # Include header row
            ws.row_dimensions[r_idx].height = 160  # Adjust the height as needed
            
        for c_idx in range(1, len(headings) + 1):
            ws.column_dimensions[chr(ord('A') + c_idx - 1)].width = 13  # Adjust the width as needed

        ws.column_dimensions[chr(ord('L'))].width = 40
        
        # Save the Excel workbook
        wb.save(file_path)
        self.parent.saveExcel_btn.setEnabled(False)

        try:
            if self.parent.finalPreparedDR['deselected'].any():
                # There are rows where 'deselected' is True
                # Save the new dataframe without the deselected datapoints
                df_filtered = self.parent.finalPreparedDR[self.parent.finalPreparedDR['deselected'] == False]
                df_filtered = df_filtered.drop(columns=['deselected'])
                df_filtered.to_excel(self.parent.pathToFinalPreparedDR_deselects, index=False)
        except:
            # There where no deselected items
            pass

        mydf = self.qtablewidget_to_dataframe()
        if len(mydf) > 0:
            self.parent.populate_load_data(mydf)
